# pylint: disable=all
"""reports.py – Excel exports, PDF marksheets, batch export, audit history export."""

import os, platform, sqlite3, subprocess
from datetime import datetime
import pandas as pd
from tkinter import filedialog, messagebox
from config import SQLITE_DB_FILE


class ReportsMixin:

    def export_results_excel(self):
        if self.results_df is None: messagebox.showerror("Error","No results"); return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")])
        if f:
            try: self.results_df.to_excel(f, index=False); messagebox.showinfo("Success",f"Exported to:\n{f}")
            except Exception as e: messagebox.showerror("Error", str(e))

    def generate_failed_report(self):
        if self.results_df is None or self.results_df.empty: messagebox.showerror("Error","No results"); return
        failed = self.results_df[self.results_df['Result']=='FAIL']
        if failed.empty: messagebox.showinfo("Info","✓ No failed students!"); return
        f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile="failed_students.xlsx")
        if f:
            try: failed.to_excel(f, index=False); messagebox.showinfo("Success",f"Exported {len(failed)} failed students to:\n{f}")
            except Exception as e: messagebox.showerror("Error", str(e))

    def generate_summary_report(self):
        messagebox.showinfo("Info","Summary PDF – coming soon!")

    def export_fixed_history(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            con = sqlite3.connect(SQLITE_DB_FILE); cur = con.cursor()
            cur.execute("SELECT fix_id,student_id,student_name,subject,field_name,"
                        "old_value,new_value,error_message,fixed_by,fixed_at FROM fixed_errors ORDER BY fixed_at DESC")
            fixes = cur.fetchall(); con.close()
            if not fixes: messagebox.showinfo("Info","No fixes to export"); return
            f = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile="fixed_errors_history.xlsx")
            if not f: return
            wb = Workbook(); ws = wb.active; ws.title = "Fixed Errors"
            hdrs = ['Fix ID','Student ID','Student Name','Subject','Field','Old Value','New Value','Error','Fixed By','Fixed At']
            ws.append(hdrs)
            fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            for cell in ws[1]:
                cell.fill = fill; cell.font = Font(bold=True, color='FFFFFF')
                cell.alignment = Alignment(horizontal='center')
            for fix in fixes: ws.append(fix)
            for col in ws.columns:
                col = list(col)
                ws.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or '')) for c in col)+2, 50)
            wb.save(f); messagebox.showinfo("Success",f"✓ Exported {len(fixes)} records to:\n{f}")
        except Exception as e: messagebox.showerror("Error", str(e))

    def generate_individual_marksheets(self):
        if self.results_df is None or self.results_df.empty: messagebox.showerror("Error","No results"); return
        out_dir = filedialog.askdirectory(title="Select output folder")
        if not out_dir: return
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors as rc
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER

            mdir = os.path.join(out_dir, "marksheets"); os.makedirs(mdir, exist_ok=True)
            self.root.config(cursor="wait"); self.root.update()
            styles = getSampleStyleSheet()
            t_sty  = ParagraphStyle('T', parent=styles['Heading1'], fontSize=24,
                                    textColor=rc.HexColor('#2C3E50'), spaceAfter=30,
                                    alignment=TA_CENTER, fontName='Helvetica-Bold')
            h_sty  = ParagraphStyle('H', parent=styles['Heading2'], fontSize=16,
                                    textColor=rc.HexColor('#34495E'), spaceAfter=12,
                                    spaceBefore=12, fontName='Helvetica-Bold')

            sid_col = next((c for c in self.results_df.columns if 'student' in c.lower() and 'id' in c.lower()), None)
            sname_c = next((c for c in self.results_df.columns if 'student' in c.lower() and 'name' in c.lower()), None)
            roll_c  = next((c for c in self.results_df.columns if 'roll' in c.lower()), None)

            if not messagebox.askyesno("Confirm", f"Generate {len(self.results_df)} marksheets?\nOutput: {mdir}"): 
                self.root.config(cursor=""); return

            generated = 0
            for _, row in self.results_df.iterrows():
                try:
                    sid   = str(row[sid_col])   if sid_col else "unknown"
                    sname = str(row[sname_c])   if sname_c else "Unknown"
                    roll  = str(row[roll_c])    if roll_c  else "N/A"
                    safe  = "".join(c for c in sid if c.isalnum() or c in '-_')
                    doc   = SimpleDocTemplate(os.path.join(mdir,f"marksheet_{safe}.pdf"), pagesize=A4)
                    els   = [Paragraph("STUDENT MARKSHEET", t_sty), Spacer(1,.3*inch)]

                    info_tbl = Table([['Student Name:',sname],['Student ID:',sid],
                                      ['Roll Number:',roll],['Academic Year:','2025-2026']],
                                     colWidths=[2*inch,4*inch])
                    info_tbl.setStyle(TableStyle([
                        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(1,0),(1,-1),'Helvetica'),
                        ('FONTSIZE',(0,0),(-1,-1),11),('BACKGROUND',(0,0),(-1,-1),rc.HexColor('#ECF0F1')),
                        ('GRID',(0,0),(-1,-1),1,rc.HexColor('#BDC3C7')),
                        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
                    els += [info_tbl, Spacer(1,.4*inch), Paragraph("MARKS OBTAINED",h_sty)]

                    marks_data = [['Subject','Obtained','Max']]
                    if sid_col and self.valid_df is not None:
                        sr = self.valid_df[self.valid_df[sid_col]==row[sid_col]]
                        sc = next((c for c in sr.columns if 'subject' in c.lower()),None)
                        mc = next((c for c in sr.columns if 'marks' in c.lower() and 'obtain' in c.lower()),None)
                        xc = next((c for c in sr.columns if 'marks' in c.lower() and 'max' in c.lower()),None)
                        for _,s in sr.iterrows():
                            marks_data.append([str(s[sc]) if sc else'–', str(s[mc]) if mc else'–', str(s[xc]) if xc else'100'])

                    m_tbl = Table(marks_data, colWidths=[3*inch,2*inch,2*inch])
                    m_tbl.setStyle(TableStyle([
                        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,0),11),
                        ('BACKGROUND',(0,0),(-1,0),rc.HexColor('#3498DB')),('TEXTCOLOR',(0,0),(-1,0),rc.whitesmoke),
                        ('ALIGN',(0,0),(-1,-1),'CENTER'),('FONTNAME',(0,1),(-1,-1),'Helvetica'),
                        ('ROWBACKGROUNDS',(0,1),(-1,-1),[rc.white,rc.HexColor('#F8F9FA')]),
                        ('GRID',(0,0),(-1,-1),1,rc.HexColor('#BDC3C7')),
                        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
                    els += [m_tbl, Spacer(1,.4*inch), Paragraph("RESULT SUMMARY",h_sty)]

                    pct  = float(str(row.get('Percentage','0%')).replace('%','') or 0)
                    res  = str(row.get('Result','FAIL'))
                    rcol = rc.HexColor('#27AE60') if res=='PASS' else rc.HexColor('#E74C3C')
                    s_tbl = Table([['Total Marks:',f"{row.get('Total',0):.0f}"],
                                   ['Percentage:',str(row.get('Percentage','0%'))],
                                   ['Grade:',str(row.get('Grade','F'))],['Result:',res]],
                                  colWidths=[2*inch,4*inch])
                    s_tbl.setStyle(TableStyle([
                        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(1,0),(1,-1),'Helvetica-Bold'),
                        ('FONTSIZE',(0,0),(-1,-1),12),('ALIGN',(0,0),(-1,-1),'LEFT'),
                        ('TEXTCOLOR',(0,3),(1,3),rcol),('BACKGROUND',(0,0),(-1,-1),rc.HexColor('#ECF0F1')),
                        ('GRID',(0,0),(-1,-1),1,rc.HexColor('#BDC3C7')),
                        ('TOPPADDING',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),10)]))
                    els.append(s_tbl)
                    doc.build(els); generated += 1
                except Exception as e: print(f"[Reports] marksheet error: {e}"); continue

            self.root.config(cursor="")
            messagebox.showinfo("Success", f"✓ {generated} marksheets generated!\nLocation: {mdir}")
            self._open_folder(mdir)
        except Exception as e:
            self.root.config(cursor=""); messagebox.showerror("Error", str(e))

    def export_all_reports(self):
        if self.results_df is None or self.results_df.empty: messagebox.showerror("Error","No results"); return
        out = filedialog.askdirectory(title="Select output folder")
        if not out: return
        try:
            rdir = os.path.join(out, f"reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}")
            os.makedirs(rdir, exist_ok=True)
            self.root.config(cursor="wait"); self.root.update()
            self.results_df.to_excel(os.path.join(rdir,"final_results.xlsx"), index=False)
            if self.error_df is not None and not self.error_df.empty:
                self.error_df.to_excel(os.path.join(rdir,"error_report.xlsx"), index=False)
            pcts = [float(str(p).replace('%','')) for p in self.results_df['Percentage'] if str(p).replace('%','').replace('.','').isdigit()]
            with open(os.path.join(rdir,"summary.txt"),'w') as f:
                f.write(f"RESULTS SUMMARY\n{'='*40}\n")
                f.write(f"Total: {len(self.results_df)}\n")
                f.write(f"Passed: {len(self.results_df[self.results_df['Result']=='PASS'])}\n")
                f.write(f"Failed: {len(self.results_df[self.results_df['Result']=='FAIL'])}\n")
                if pcts: f.write(f"Avg: {sum(pcts)/len(pcts):.2f}%  High: {max(pcts):.2f}%  Low: {min(pcts):.2f}%\n")
            self.root.config(cursor="")
            messagebox.showinfo("Success", f"✓ All reports exported!\nLocation: {rdir}")
            self._open_folder(rdir)
        except Exception as e:
            self.root.config(cursor=""); messagebox.showerror("Error", str(e))

    def _open_folder(self, path):
        try:
            if platform.system()=='Windows': os.startfile(path)
            elif platform.system()=='Darwin': subprocess.run(['open',path],check=False)
            else: subprocess.run(['xdg-open',path],check=False)
        except Exception: pass
